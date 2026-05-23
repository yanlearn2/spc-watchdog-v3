#!/usr/bin/env python3
"""
中控检测数据实时监控报警器 V3.0 — 数据库版

核心能力：
  · 自动识别标准行，从标准行自动提取规格
  · 白名单机制：只监控重点样品，其余跳过
  · 数据库同步：样品数据、报警记录、规格全部入库
  · 数据库后端可插拔（当前: SQLite，预留 MySQL/PostgreSQL 接口）
  · 自动判断 "(参考)" → monitor 模式（只记日志、不报警）
  · 无 "(参考)" → alarm 模式（发飞书通知）
  · 正则匹配样品名称（标准行写 `一浸液[GHIJKL]槽`，匹配数据行 `一浸液G槽`）
  · 自动跳过 "实测"、"NotNull"、"IsNull" 等非量化规格

用法：
  python3 watchdog.py                  # 持续监控
  python3 watchdog.py --once           # 只跑一次
  python3 watchdog.py --init-db        # 只初始化数据库（不扫描）
  python3 watchdog.py --daemon         # 后台运行
"""
import json, os, sys, time, re, logging, hashlib, sqlite3, signal, random
from datetime import datetime
from pathlib import Path
from abc import ABC, abstractmethod
from logging.handlers import RotatingFileHandler

try:
    import openpyxl
except ImportError:
    print("缺少 openpyxl，请执行: pip install openpyxl"); sys.exit(1)
try:
    import requests
except ImportError:
    print("缺少 requests，请执行: pip install requests"); sys.exit(1)

# ─── 路径 ──────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent
CONFIG_PATH = SCRIPT_DIR / "config.json"
LOG_PATH = SCRIPT_DIR / "watchdog.log"
STOP_FLAG = False  # 优雅退出标志

# ─── 日志配置（基础） ──────────────────────────────────
logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[logging.StreamHandler()])
log = logging.getLogger("watchdog")

# ─── 信号处理 ──────────────────────────────────────────
def signal_handler(signum, frame):
    global STOP_FLAG
    log.warning(f"🛑 收到信号 {signum}，优雅退出中...")
    STOP_FLAG = True

# ─── 配置 ──────────────────────────────────────────────
def load_config():
    if not CONFIG_PATH.exists():
        log.error(f"❌ 配置文件不存在: {CONFIG_PATH}"); sys.exit(1)
    with open(CONFIG_PATH, encoding="utf-8") as f:
        return json.load(f)

CONFIG = load_config()

# ─── 信号注册 ──────────────────────────────────────────
signal.signal(signal.SIGINT, signal_handler)
signal.signal(signal.SIGTERM, signal_handler)

# ─── 日志文件(滚动) ──────────────────────────────────
log_cfg = CONFIG.get("log", {})
max_mb = log_cfg.get("max_file_size_mb", 50)
backup_count = log_cfg.get("backup_count", 3)
log_level = getattr(logging, log_cfg.get("level", "INFO").upper(), logging.INFO)
log.setLevel(log_level)
file_handler = RotatingFileHandler(LOG_PATH, maxBytes=max_mb * 1024 * 1024, backupCount=backup_count, encoding="utf-8")
file_handler.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
log.addHandler(file_handler)

# ─── 进度记录 ──────────────────────────────────────────
PROGRESS_FILE = Path(CONFIG["watch"]["progress_file"])
if not PROGRESS_FILE.is_absolute():
    PROGRESS_FILE = SCRIPT_DIR / PROGRESS_FILE

def load_progress():
    if PROGRESS_FILE.exists():
        with open(PROGRESS_FILE) as f:
            return json.load(f)
    return {"last_mtime": 0}

def save_progress(p):
    with open(PROGRESS_FILE, "w") as f:
        json.dump(p, f, ensure_ascii=False, indent=2)


# ════════════════════════════════════════════════════════════
# 数据库后端抽象接口
# ════════════════════════════════════════════════════════════
#
# 架构说明：
#   DatabaseBackend 是所有数据库后端的抽象基类。
#   watchdog 中所有数据库操作均通过此接口完成，
#   更换数据库只需新增一个实现类 + 在 _create_backend() 中注册。
#
# 添加新数据库后端的步骤：
#   1. 新建一个类继承 DatabaseBackend
#   2. 实现所有 @abstractmethod
#   3. 在 _create_backend() 中 if db_type == "mysql": return MySQLBackend()
#   4. 在 config.json 中 database.type 填写对应名称
#   5. 在 config.json 的 database 中填写该后端所需的配置参数
#
# 当前实现：
#   - SQLiteBackend（默认，零配置）
#   - (预留) MySQLBackend
#   - (预留) PostgreSQLBackend
#   - (预留) FeishuBitableBackend
#
# ════════════════════════════════════════════════════════════

class DatabaseBackend(ABC):
    """数据库后端抽象接口 — 所有数据库操作均通过此接口"""

    # ─── 连接管理 ──────────────────────────────────────
    @abstractmethod
    def connect(self, db_config: dict):
        """建立数据库连接
        db_config: config.json 中 database 段的完整配置
        实现示例 (SQLite): sqlite3.connect(db_config.get("sqlite_path", "./检测数据库.sqlite"))
        实现示例 (MySQL):  pymysql.connect(host=..., user=..., password=..., database=...)
        """

    @abstractmethod
    def close(self):
        """关闭数据库连接"""

    @abstractmethod
    def is_connected(self) -> bool:
        """返回连接是否有效"""

    # ─── 建表 ──────────────────────────────────────────
    @abstractmethod
    def init_schema(self):
        """创建所需的表结构（若不存在）
        需要创建的表（名称可自定义，但以下含义必须保留）：
          - samples:       原始检测数据行
          - alarm_log:     报警记录
          - qc_specs:      QC规格表（报警规则的权威来源）
          - sync_log:      同步日志（记录每次同步的批次信息）
        """

    # ─── 样品数据 ──────────────────────────────────────
    @abstractmethod
    def insert_sample(self, row_hash: str, sample_name: str, test_date: str,
                      workshop: str, data_json: str, source_row: int,
                      source_file: str, sheet_name: str, synced_at: str):
        """插入一条样品检测记录，返回自增ID（0表示已存在/失败）"""

    @abstractmethod
    def sample_exists(self, row_hash: str) -> bool:
        """检查该行指纹是否已存在"""

    @abstractmethod
    def get_sample_count(self) -> int:
        """返回 samples 表中的总记录数"""

    # ─── 报警记录 ──────────────────────────────────────
    @abstractmethod
    def insert_alarm(self, sample_id: int, sample_name: str, item: str,
                     value: float, spec: str, unit: str, deviation: str,
                     mode: str, sheet: str, alarmed_at: str):
        """插入一条报警记录，返回自增ID"""

    @abstractmethod
    def get_alarm_count(self, since: str = "") -> int:
        """返回报警记录数；since 为 ISO 时间字符串，如 '2026-05-21 00:00:00'"""

    # ─── 规格管理 ──────────────────────────────────────
    @abstractmethod
    def upsert_spec(self, sample_pattern: str, item: str, upper: float,
                    lower: float, mode: str, unit: str, sheet: str):
        """插入或更新一条QC规格（复合主键: sample_pattern + item）"""

    @abstractmethod
    def get_specs(self, sample_filter: str = "") -> list:
        """获取QC规格列表
        返回: [{"sample_pattern": ..., "item": ..., "upper": ..., "lower": ..., "mode": ..., "unit": ..., "sheet": ...}, ...]
        若 sample_filter 非空，仅返回 sample_pattern 匹配该条件的记录
        """

    @abstractmethod
    def delete_spec(self, sample_pattern: str, item: str):
        """删除一条QC规格"""

    # ─── 同步日志 ──────────────────────────────────────
    @abstractmethod
    def log_sync(self, sheet_name: str, total_rows: int, new_rows: int,
                 alarms_count: int, synced_at: str):
        """记录一次同步日志"""

    @abstractmethod
    def get_sync_log(self, limit: int = 10) -> list:
        """获取最近的同步日志"""

    # ─── 批量操作（性能优化） ──────────────────────────
    @abstractmethod
    def insert_samples_batch(self, rows: list) -> int:
        """批量插入样品记录，返回新增数量
        rows: [(row_hash, sample_name, test_date, workshop, data_json, source_row, source_file, sheet_name, synced_at), ...]
        """

    @abstractmethod
    def insert_alarms_batch(self, alarms: list) -> int:
        """批量插入报警记录
        alarms: [(sample_id, sample_name, item, value, spec, unit, deviation, mode, sheet, alarmed_at), ...]
        """

    @abstractmethod
    def get_all_existing_hashes(self) -> dict:
        """获取所有已存在的自然键→row_hash映射，用于内存去重+更新检测
        返回: {(sample_name, test_date): row_hash}
        """

    @abstractmethod
    def get_specs_dict(self) -> dict:
        """从数据库加载规格，返回 {sheet_name: {sample_regex: [(item, upper, lower, mode, unit), ...]}}
        替代每次从标准行解析，大幅减少repeat工作量
        """


class SQLiteBackend(DatabaseBackend):
    """SQLite 数据库后端 — 零配置，默认使用"""

    def __init__(self):
        self.conn = None

    def connect(self, db_config: dict):
        db_path = db_config.get("sqlite_path", "./检测数据库.sqlite")
        if not os.path.isabs(db_path):
            db_path = str(SCRIPT_DIR / db_path)
        self.conn = sqlite3.connect(db_path)
        self.conn.execute("PRAGMA journal_mode=WAL")
        self.conn.execute("PRAGMA foreign_keys=ON")
        log.info(f"📁 数据库: {db_path}")

        # 验证连接
        if hasattr(self, 'conn') and self.conn:
            try:
                self.conn.execute("SELECT 1").fetchone()
            except Exception:
                self.conn = None

    def close(self):
        if self.conn:
            self.conn.close()
            self.conn = None

    def is_connected(self) -> bool:
        return self.conn is not None

    def init_schema(self):
        self.conn.executescript("""
            -- 样品检测记录（当年热数据）
            CREATE TABLE IF NOT EXISTS samples (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                row_hash    TEXT UNIQUE,                    -- Excel行指纹(MD5)，用于去重
                sample_name TEXT NOT NULL,                  -- 样品名称
                test_date   TEXT,                            -- 检测日期
                workshop    TEXT,                            -- 车间/工作表名
                source_file TEXT,                            -- 来源文件名
                source_row  INTEGER,                         -- 来源Excel行号
                sheet_name  TEXT,                            -- 工作表名
                data_json   TEXT,                            -- 全部检测结果 {"Co":xx, "Ni":xx, ...}
                synced_at   DATETIME,                        -- 同步时间
                created_at  DATETIME DEFAULT (datetime('now','localtime'))
            );
            CREATE INDEX IF NOT EXISTS idx_samples_date  ON samples(test_date);
            CREATE INDEX IF NOT EXISTS idx_samples_name  ON samples(sample_name);
            CREATE INDEX IF NOT EXISTS idx_samples_workshop ON samples(workshop);
            CREATE INDEX IF NOT EXISTS idx_samples_hash  ON samples(row_hash);

            -- 报警记录表（支持闭环跟踪）
            CREATE TABLE IF NOT EXISTS alarm_log (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                sample_name TEXT,                            -- 样品名称
                item        TEXT,                             -- 超限的检测项目
                value       REAL,                             -- 检测值
                spec        TEXT,                             -- 规格描述
                unit        TEXT,                             -- 单位
                deviation   TEXT,                             -- 偏差百分比
                mode        TEXT DEFAULT 'alarm',             -- alarm | monitor
                sheet       TEXT,                             -- 来源工作表
                alarm_mode  TEXT DEFAULT 'alarm',             -- alarm | monitor
                sample_id   INTEGER,                          -- 关联 samples.id
                level       TEXT DEFAULT '三级',              -- 一级/二级/三级
                status      TEXT DEFAULT '待处置',            -- 待处置/已处置/已闭环
                reaction    TEXT DEFAULT '',                  -- 处置措施
                alarmed_at  DATETIME,                        -- 报警时间
                closed_at   DATETIME,                         -- 闭环时间
                created_at  DATETIME DEFAULT (datetime('now','localtime'))
            );
            CREATE INDEX IF NOT EXISTS idx_alarm_status  ON alarm_log(status);
            CREATE INDEX IF NOT EXISTS idx_alarm_time   ON alarm_log(alarmed_at);
            CREATE INDEX IF NOT EXISTS idx_alarm_mode   ON alarm_log(mode);

            -- QC规格表（从标准行自动提取，充当报警规则的权威来源）
            CREATE TABLE IF NOT EXISTS qc_specs (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                sample_pattern  TEXT NOT NULL,                -- 样品名称正则
                item            TEXT NOT NULL,                -- 检测项目
                upper_limit     REAL,                         -- 上限
                lower_limit     REAL,                         -- 下限
                mode            TEXT DEFAULT 'alarm',         -- alarm | monitor
                unit            TEXT DEFAULT '',              -- 单位
                sheet           TEXT DEFAULT '',              -- 来源工作表
                source          TEXT DEFAULT 'auto',          -- auto(自动提取) | manual(手动录入)
                updated_at      DATETIME DEFAULT (datetime('now','localtime')),
                UNIQUE(sample_pattern, item)
            );
            CREATE INDEX IF NOT EXISTS idx_qc_specs_pattern ON qc_specs(sample_pattern);

            -- 同步日志（记录每次扫描批次）
            CREATE TABLE IF NOT EXISTS sync_log (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                sheet_name  TEXT,
                total_rows  INTEGER DEFAULT 0,
                new_rows    INTEGER DEFAULT 0,
                alarms_count INTEGER DEFAULT 0,
                synced_at   DATETIME,
                created_at  DATETIME DEFAULT (datetime('now','localtime'))
            );
        """)
        self.conn.commit()

    def insert_sample(self, row_hash, sample_name, test_date, workshop,
                      data_json, source_row, source_file, sheet_name, synced_at):
        try:
            cur = self.conn.execute(
                """INSERT INTO samples
                   (row_hash, sample_name, test_date, workshop, data_json, source_row, source_file, sheet_name, synced_at)
                   VALUES (?,?,?,?,?,?,?,?,?)""",
                (row_hash, sample_name, test_date, workshop, data_json,
                 source_row, source_file, sheet_name, synced_at)
            )
            self.conn.commit()
            return cur.lastrowid
        except sqlite3.IntegrityError:
            return 0

    def sample_exists(self, row_hash):
        return bool(self.conn.execute(
            "SELECT 1 FROM samples WHERE row_hash=?", (row_hash,)
        ).fetchone())

    def get_sample_count(self):
        return self.conn.execute("SELECT COUNT(*) FROM samples").fetchone()[0]

    def insert_alarm(self, sample_id, sample_name, item, value, spec,
                     unit, deviation, mode, sheet, alarmed_at):
        cur = self.conn.execute(
            """INSERT INTO alarm_log
               (sample_id, sample_name, item, value, spec, unit, deviation, mode, sheet, alarm_mode, alarmed_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (sample_id, sample_name, item, value, spec, unit, deviation, mode, sheet, mode, alarmed_at)
        )
        self.conn.commit()
        return cur.lastrowid

    def get_alarm_count(self, since=""):
        if since:
            return self.conn.execute(
                "SELECT COUNT(*) FROM alarm_log WHERE alarmed_at >= ?", (since,)
            ).fetchone()[0]
        return self.conn.execute("SELECT COUNT(*) FROM alarm_log").fetchone()[0]

    def upsert_spec(self, sample_pattern, item, upper, lower, mode, unit, sheet):
        self.conn.execute(
            """INSERT OR REPLACE INTO qc_specs
               (sample_pattern, item, upper_limit, lower_limit, mode, unit, sheet, updated_at)
               VALUES (?,?,?,?,?,?,?,datetime('now','localtime'))""",
            (sample_pattern, item, upper, lower, mode, unit, sheet)
        )
        self.conn.commit()

    def get_specs(self, sample_filter=""):
        if sample_filter:
            rows = self.conn.execute(
                "SELECT sample_pattern, item, upper_limit, lower_limit, mode, unit, sheet FROM qc_specs WHERE sample_pattern LIKE ? ORDER BY sample_pattern, item",
                (f"%{sample_filter}%",)
            ).fetchall()
        else:
            rows = self.conn.execute(
                "SELECT sample_pattern, item, upper_limit, lower_limit, mode, unit, sheet FROM qc_specs ORDER BY sample_pattern, item"
            ).fetchall()
        return [
            {"sample_pattern": r[0], "item": r[1], "upper": r[2], "lower": r[3],
             "mode": r[4], "unit": r[5], "sheet": r[6]}
            for r in rows
        ]

    def delete_spec(self, sample_pattern, item):
        self.conn.execute(
            "DELETE FROM qc_specs WHERE sample_pattern=? AND item=?",
            (sample_pattern, item)
        )
        self.conn.commit()

    def log_sync(self, sheet_name, total_rows, new_rows, alarms_count, synced_at):
        self.conn.execute(
            "INSERT INTO sync_log (sheet_name, total_rows, new_rows, alarms_count, synced_at) VALUES (?,?,?,?,?)",
            (sheet_name, total_rows, new_rows, alarms_count, synced_at)
        )
        self.conn.commit()

    def get_sync_log(self, limit=10):
        rows = self.conn.execute(
            "SELECT sheet_name, total_rows, new_rows, alarms_count, synced_at, created_at FROM sync_log ORDER BY id DESC LIMIT ?",
            (limit,)
        ).fetchall()
        return [
            {"sheet": r[0], "total": r[1], "new": r[2], "alarms": r[3],
             "synced_at": r[4], "created_at": r[5]}
            for r in rows
        ]

    # ─── 批量操作 ──────────────────────────────────────
    def insert_samples_batch(self, rows):
        if not rows:
            return 0
        try:
            self.conn.executemany(
                """INSERT OR IGNORE INTO samples
                   (row_hash, sample_name, test_date, workshop, data_json, source_row, source_file, sheet_name, synced_at)
                   VALUES (?,?,?,?,?,?,?,?,?)""",
                rows
            )
            self.conn.commit()
            return len(rows)
        except Exception:
            return 0

    def insert_alarms_batch(self, alarms):
        if not alarms:
            return 0
        try:
            self.conn.executemany(
                """INSERT INTO alarm_log
                   (sample_id, sample_name, item, value, spec, unit, deviation, mode, sheet, alarm_mode, alarmed_at)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                alarms
            )
            self.conn.commit()
            return len(alarms)
        except Exception:
            return 0

    def get_all_existing_hashes(self):
        """获取所有已存在的 row_hash 和自然键，返回 {natural_key: row_hash}
        natural_key = (sample_name, test_date)
        """
        rows = self.conn.execute("SELECT row_hash, sample_name, test_date FROM samples").fetchall()
        result = {}
        for r in rows:
            key = (r[1], r[2])  # (sample_name, test_date)
            result[key] = r[0]  # row_hash
        return result

    def get_specs_dict(self):
        """从数据库加载规格，返回 {sheet_name: {sample_regex: [(item, upper, lower, mode, unit), ...]}}"""
        rows = self.conn.execute(
            "SELECT sheet, sample_pattern, item, upper_limit, lower_limit, mode, unit FROM qc_specs ORDER BY sheet, sample_pattern, item"
        ).fetchall()
        result = {}
        for r in rows:
            sheet, pattern, item, upper, lower, mode, unit = r
            if sheet not in result:
                result[sheet] = {}
            if pattern not in result[sheet]:
                result[sheet][pattern] = []
            result[sheet][pattern].append((item, upper, lower, mode, unit))
        return result


# ─── 数据库后端工厂 ──────────────────────────────────────
def create_backend(db_config: dict) -> DatabaseBackend:
    """工厂方法：根据配置创建对应的数据库后端实例"""
    db_type = db_config.get("type", "sqlite").lower()

    if db_type == "sqlite":
        import sqlite3
        backend = SQLiteBackend()
    # ────────────────────────────────────────────────────
    # 扩展点：在此添加新的数据库后端
    # 示例：
    # elif db_type == "mysql":
    #     import pymysql
    #     backend = MySQLBackend()
    # elif db_type == "postgresql":
    #     import psycopg2
    #     backend = PostgreSQLBackend()
    # elif db_type == "feishu_bitable":
    #     backend = FeishuBitableBackend()
    # ────────────────────────────────────────────────────
    else:
        log.warning(f"⚠ 未知数据库类型: {db_type}，回退到 SQLite")
        import sqlite3
        backend = SQLiteBackend()

    backend.connect(db_config)
    return backend


# ─── 初始化数据库 ──────────────────────────────────────
def init_database():
    """初始化数据库（创建表结构）"""
    db_config = CONFIG.get("database", {"type": "sqlite", "sqlite_path": "./检测数据库.sqlite"})
    db = create_backend(db_config)

    if not db.is_connected():
        log.error("❌ 无法连接数据库，请检查 config.json 中的 database 配置")
        sys.exit(1)

    db.init_schema()

    # 输出初始化信息
    sample_count = db.get_sample_count()
    alarm_count = db.get_alarm_count()
    specs = db.get_specs()
    sync_logs = db.get_sync_log(5)
    db.close()

    print(f"✅ 数据库初始化完成")
    print(f"📊 当前状态:")
    print(f"   · samples 表: {sample_count} 条样品记录")
    print(f"   · alarm_log 表: {alarm_count} 条报警记录")
    print(f"   · qc_specs 表: {len(specs)} 条QC规格")
    if sync_logs:
        print(f"   · sync_log 表: 最近同步记录:")
        for sl in sync_logs[:3]:
            print(f"     - {sl['sheet']}: {sl['new']} 新增 / {sl['alarms']} 报警 ({sl['synced_at']})")


# ─── 进度记录 ──────────────────────────────────────────

# ─── 规格值解析（核心） ──────────────────────────────
def parse_spec_value(raw):
    """
    解析标准行中的规格值，返回 (upper, lower, mode, unit)

    输入示例: '≤10g/L(参考)'   '≥80'   '80~110'   '0.005~0.015(参考)'   '实测'   'NotNull'
    """
    if raw is None:
        return None, None, "skip", ""
    s = str(raw).strip()
    if not s or s == "实测" or s == "IsNull":
        return None, None, "skip", ""
    if s == "NotNull":
        return None, None, "notnull", ""

    # 提取单位（在数值后面的字母）
    unit = ""
    unit_match = re.search(r'[a-zA-Z/μ%]+', s)
    if unit_match:
        unit = unit_match.group()

    # 判断是否为参考项（带参考字样的不报警）
    is_ref = bool(re.search(r'参考', s))
    mode = "monitor" if is_ref else "alarm"

    # Clean: 去掉括号内文字（如(参考)、(参考2)、(自检)），再去掉单位字符
    clean = re.sub(r'[（(][^）)]*[）)]', '', s)  # 删括号及内容
    clean = re.sub(r'[a-zA-Z/μ%（）()\\s]+', '', clean).strip()  # 删单位字符

    # ≤X 或 <X
    m = re.match(r'[≤<]\s*([\d.]+)', clean)
    if m:
        return float(m.group(1)), None, mode, unit

    # ≥X 或 >X
    m = re.match(r'[≥>]\s*([\d.]+)', clean)
    if m:
        return None, float(m.group(1)), mode, unit

    # X~Y 或 X-Y 或 X～Y（范围）
    m = re.match(r'([\d.]+)\s*[~－\u2013-]\s*([\d.]+)', clean)
    if m:
        lower, upper = float(m.group(1)), float(m.group(2))
        return upper, lower, mode, unit

    # 纯数字（如 "0.005" — 无比较符，通常表示上限）
    m = re.match(r'^([\d.]+)$', clean)
    if m:
        return float(m.group(1)), None, mode, unit

    return None, None, "skip", ""


def extract_specs_from_standard_rows(ws, sheet_name, whitelist_patterns=None):
    """
    从标准行提取规格，返回:
    { sample_regex: [(item_name, upper, lower, mode, unit), ...] }

    若 whitelist_patterns 不为空，只提取匹配白名单的标准行规格
    """
    standard_rows = []
    headers = {}

    # 用 iter_rows 逐行读取前50行（read_only模式只能用迭代）
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True), 1):
        if not row:
            continue
        cell_a = str(row[0]).strip() if row[0] else ""
        cell_f = str(row[5]).strip() if len(row) > 5 and row[5] else ""

        # 第一行：找表头
        if r_idx == 1:
            for c, v in enumerate(row, 1):
                if v:
                    headers[c] = str(v).strip()
            continue

        # 标准行特征
        is_standard = ("标准" in cell_a or "无标准" in cell_a)
        # 检测是否为正则表达式（含 [ ] 或 .* 或 ^ $ 等特殊字符）
        has_regex = any(ch in cell_f for ch in '[]*+?^$.|\\') and not cell_f.isalnum()

        if is_standard or (has_regex and cell_f and r_idx <= 20):
            standard_rows.append((r_idx, cell_f, row))

    if not standard_rows or not headers:
        return {}, None

    # 缓存表头供 scan_sheet 使用
    ws._cached_headers = headers

    # 遍历标准行提取规格
    specs = {}
    for row_idx, sample_regex, row in standard_rows:
        # 白名单过滤：若白名单有效，只保留匹配的类别
        if not match_whitelist(sample_regex):
            continue

        items = []
        for col_idx, col_name in headers.items():
            if col_idx <= 6:
                continue

            raw = row[col_idx - 1] if col_idx - 1 < len(row) else None
            upper, lower, mode, unit = parse_spec_value(raw)

            if mode == "skip":
                continue

            items.append((col_name, upper, lower, mode, unit))

        if items:
            specs[sample_regex] = items

    return specs, headers


def match_sample_to_spec(sample_name, specs):
    """用预编译正则匹配样品名称到规格"""
    for regex_str, items in specs.items():
        pat = get_spec_regex(regex_str)
        if pat is not None and pat.search(sample_name):
            return items
        if pat is None:
            # 正则无效，模糊匹配
            simplified = regex_str.replace("[", "").replace("]", "").replace(".*", "")
            if simplified in sample_name or sample_name.startswith(simplified):
                return items
    return []


def check_row_against_specs(row_data, sample_name, specs, alarm_mode_config):
    """检查一行数据是否符合规格，返回报警列表"""
    items = match_sample_to_spec(sample_name, specs)
    if not items:
        return []

    hits = []
    for col_name, upper, lower, base_mode, unit in items:
        # 应用规格覆盖（保留原始规格用于显示，用 alarm_阈值 判断是否报警）
        display_upper, display_lower, alarm_upper, alarm_lower, unit, has_override = \
            apply_spec_overrides(sample_name, col_name, upper, lower, unit)
        # 列名到值
        raw_val = row_data.get(col_name)
        if raw_val is None or str(raw_val).strip() == "":
            continue

        try:
            value = float(raw_val)
        except (ValueError, TypeError):
            continue

        # 检查超限（用 alarm_阈值 判断，用 display_规格 显示）
        violated = False
        spec_desc = ""
        deviation = ""

        if display_upper is not None and display_lower is not None:
            spec_desc = f"{display_lower}~{display_upper}"
            if alarm_upper is not None and value > alarm_upper:
                violated = True
                deviation = f"+{(value/alarm_upper-1)*100:.0f}%" if alarm_upper > 0 else f"+{(value-display_lower):.4f}"
            elif alarm_lower is not None and value < alarm_lower:
                violated = True
                deviation = f"-{(1-value/alarm_lower)*100:.0f}%" if alarm_lower > 0 else f"-{alarm_lower-value:.4f}"
        elif display_upper is not None:
            spec_desc = f"≤{display_upper}"
            if alarm_upper is not None and value > alarm_upper:
                violated = True
                deviation = f"+{(value/alarm_upper-1)*100:.0f}%" if alarm_upper > 0 else f"+{value:.4f}"
        elif display_lower is not None:
            spec_desc = f"≥{display_lower}"
            if alarm_lower is not None and value < alarm_lower:
                violated = True
                deviation = f"-{(1-value/alarm_lower)*100:.0f}%" if alarm_lower > 0 else f"-{alarm_lower-value:.4f}"

        # 如果用了覆盖阈值，在规格描述中注明
        if has_override and violated:
            spec_desc += f"（报警≥{alarm_upper or alarm_lower}）"

        if not violated:
            continue

        # 决定模式：先看用户覆盖，再看标准行判断
        override_key = f"{sample_name}/{col_name}"
        overrides = alarm_mode_config.get("overrides", {})
        if override_key in overrides:
            mode = overrides[override_key]
        else:
            mode = base_mode

        # 提取检测时间（从行数据中取"送样日期"或"送样时间"）
        alarm_time = ""
        for time_key in ["送样日期", "送样时间", "检测日期", "报出时间"]:
            if time_key in row_data:
                alarm_time = str(row_data[time_key])
                break

        hits.append({
            "sample": sample_name,
            "item": col_name,
            "value": value,
            "unit": unit,
            "spec": spec_desc,
            "deviation": deviation,
            "mode": mode,       # alarm | monitor
            "sheet": row_data.get("_sheet", ""),
            "time": alarm_time  # 送样日期/时间
        })

    return hits


def apply_spec_overrides(sample_name, item, upper, lower, unit):
    """
    应用 config.json 中的 spec_overrides。
    返回 (display_upper, display_lower, alarm_upper, alarm_lower, unit, has_override)
    
    - display_upper/display_lower: 原始规格（用于显示）
    - alarm_upper/alarm_lower: 报警阈值（用于判断是否报警，默认等于原始值）
    - has_override: 是否被覆盖
    """
    overrides = CONFIG.get("spec_overrides", {})
    if not overrides:
        return upper, lower, upper, lower, unit, False

    for key, spec in overrides.items():
        pattern_part, item_part = key.split("/", 1)
        try:
            matched = re.search(pattern_part, sample_name)
        except re.error:
            matched = pattern_part in sample_name

        if matched and item_part == item:
            alarm_upper = spec.get("alarm_upper", upper)
            alarm_lower = spec.get("alarm_lower", lower)
            new_unit = spec.get("unit", unit)
            return upper, lower, alarm_upper, alarm_lower, new_unit, True

    return upper, lower, upper, lower, unit, False


# ─── 重试装饰器 ────────────────────────────────────────
def retry_on_failure(max_attempts=3, base_delay=5, max_delay=60):
    """带指数退避的重试装饰器"""
    def decorator(func):
        def wrapper(*args, **kwargs):
            watch_cfg = CONFIG.get("watch", {})
            retry_cfg = watch_cfg.get("retry", {})
            attempts = retry_cfg.get("max_attempts", max_attempts)
            base = retry_cfg.get("base_delay_seconds", base_delay)
            max_d = retry_cfg.get("max_delay_seconds", max_delay)
            last_exc = None
            for attempt in range(1, attempts + 1):
                try:
                    return func(*args, **kwargs)
                except (PermissionError, OSError) as e:
                    last_exc = e
                    if attempt < attempts:
                        delay = min(base * 2 ** (attempt - 1) + random.uniform(0, 1), max_d)
                        log.warning(f"🔄 第{attempt}次打开失败: {e}，{delay:.0f}s后重试...")
                        time.sleep(delay)
                    else:
                        log.error(f"❌ 重试{attempts}次均失败: {e}")
            raise last_exc
        return wrapper
    return decorator


# ─── 解密Excel（支持密码+重试） ──────────────────────
@retry_on_failure()
def open_excel(path, password=None):
    """打开Excel，支持密码。返回 openpyxl.Workbook（带重试）"""
    if password:
        try:
            import msoffcrypto
            with open(path, "rb") as f:
                office = msoffcrypto.OfficeFile(f)
                office.load_key(password=password)
                import tempfile
                tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
                office.decrypt(tmp)
                tmp.close()
            wb = openpyxl.load_workbook(tmp.name, data_only=True, read_only=True)
            os.unlink(tmp.name)
            return wb
        except ImportError:
            log.warning("⚠ 需要密码才能打开Excel，但 msoffcrypto-tool 未安装")
            log.warning("  执行: pip install msoffcrypto-tool")
            return None
        except Exception as e:
            log.warning(f"⚠ 解密失败: {e}，尝试直接打开...")

    if os.path.exists(path):
        return openpyxl.load_workbook(path, data_only=True, read_only=True)
    return None


# ─── 正则缓存（预编译提升性能） ──────────────────────
_WHITELIST_COMPILED = None  # [re.Pattern, ...] or None
_SPEC_REGEX_CACHE = {}      # {regex_str: re.Pattern}

def compile_whitelist():
    """预编译白名单正则表达式"""
    global _WHITELIST_COMPILED
    filter_config = CONFIG.get("filter", {})
    enabled = filter_config.get("whitelist_enabled", False)
    if not enabled:
        _WHITELIST_COMPILED = None
        return None
    patterns = filter_config.get("whitelist_patterns", [])
    if not patterns:
        _WHITELIST_COMPILED = None
        return None
    compiled = []
    for wp in patterns:
        try:
            compiled.append(re.compile(wp))
        except re.error:
            compiled.append(None)
    _WHITELIST_COMPILED = compiled
    return compiled

def match_whitelist(sample_name):
    """用预编译白名单判断样品名是否匹配"""
    global _WHITELIST_COMPILED
    if _WHITELIST_COMPILED is None:
        return True
    for pat in _WHITELIST_COMPILED:
        if pat is None:
            continue
        if pat.search(sample_name):
            return True
    return False

def get_spec_regex(regex_str):
    """获取/缓存编译后的规格正则"""
    if regex_str not in _SPEC_REGEX_CACHE:
        try:
            _SPEC_REGEX_CACHE[regex_str] = re.compile(regex_str)
        except re.error:
            _SPEC_REGEX_CACHE[regex_str] = None
    return _SPEC_REGEX_CACHE[regex_str]

# ─── 白名单过滤 ──────────────────────────────────────
def get_whitelist():
    """获取白名单模式列表（如果启用）（兼容旧接口，返回原始字符串列表）"""
    filter_config = CONFIG.get("filter", {})
    enabled = filter_config.get("whitelist_enabled", False)
    if not enabled:
        return None  # None = 不过滤
    patterns = filter_config.get("whitelist_patterns", [])
    if not patterns:
        return None
    return patterns


def sheet_has_whitelist_match(ws, sheet_name, whitelist_patterns):
    """
    快速检查该工作表是否有匹配白名单的数据行
    通过扫描标准行（R2~R6）快速判断
    """
    for row in ws.iter_rows(min_row=2, max_row=10, values_only=True):
        if not row or len(row) < 6:
            continue
        sample_name = str(row[5]).strip() if row[5] else ""
        if not sample_name or sample_name == "None":
            continue
        if match_whitelist(sample_name):
            return True
    return False


# ─── 扫描一个工作表 ────────────────────────────────────
def scan_sheet(ws, sheet_name, specs_cache, whitelist_patterns, db, existing_hashes=None):
    """扫描一个工作表，返回 (数据行数, 报警列表, 新增行数, 耗时秒)

    优化项：
    - 批量写入(500行/flush)替代逐行INSERT
    - 内存hash缓存替代逐行DB查询
    - 规格从qc_specs表缓存加载（跳过标准行重解析）
    - 移除max_row调用（read_only下触发全表扫描）
    """
    t0 = time.time()
    all_alarms = []
    data_count = 0
    new_rows = 0
    empty_count = 0
    max_empty = 100
    BATCH_SIZE = 500

    # 1. 提取/加载规格
    sheet_specs = specs_cache.get(sheet_name)
    if sheet_specs is None:
        sheet_specs, headers = extract_specs_from_standard_rows(ws, sheet_name, whitelist_patterns)
        specs_cache[sheet_name] = sheet_specs
        if sheet_specs:
            log.info(f"  📋 {sheet_name}: 提取 {len(sheet_specs)} 条规格")
            for regex, items in list(sheet_specs.items())[:5]:
                items_desc = " | ".join(f"{i[0]}:{i[3]}" for i in items[:4])
                log.info(f"     · {regex[:30]:30s} → {items_desc}")
            if len(sheet_specs) > 5:
                log.info(f"     ... 共 {len(sheet_specs)} 条规格")
            # 写入QC规格到数据库
            for regex_str, items in sheet_specs.items():
                for item_name, upper, lower, mode, unit in items:
                    db.upsert_spec(regex_str, item_name, upper, lower, mode, unit, sheet_name)
            log.info(f"     💾 已写入 {sum(len(v) for v in sheet_specs.values())} 条规格到数据库")

    if not sheet_specs:
        return 0, [], 0, time.time() - t0

    # 2. 读取表头
    headers = getattr(ws, '_cached_headers', {})
    if not headers:
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            for c, v in enumerate(row, 1):
                if v:
                    headers[c] = str(v).strip()
            break
        ws._cached_headers = headers

    # 3. 确定日期列
    date_col_idx = None
    for col_idx, col_name in headers.items():
        if col_name and '日期' in str(col_name):
            date_col_idx = col_idx
            break

    # 4. 扫描数据行
    synced_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    excel_filename = os.path.basename(CONFIG["excel"]["path"])

    # 批量缓冲区
    sample_batch = []
    alarm_batch = []
    # 初始空hash集合
    if existing_hashes is None:
        existing_hashes = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row[:10]):
            empty_count += 1
            if empty_count > max_empty:
                break
            continue
        empty_count = 0

        cell_a = str(row[0]).strip() if row[0] else ""
        cell_f = str(row[5]).strip() if len(row) > 5 and row[5] else ""

        # 跳过标准行
        if "标准" in cell_a or "无标准" in cell_a:
            continue
        if not cell_f or cell_f == "None":
            continue
        if re.search(r'[.*\\[\\]]', cell_f):
            continue

        sample_name = cell_f

        # 白名单过滤
        if not match_whitelist(sample_name):
            continue

        # 组装行数据
        row_data = {"_sheet": sheet_name}
        for col_idx, val in enumerate(row, 1):
            if val is not None and col_idx in headers:
                if isinstance(val, (str, int, float, bool)):
                    row_data[headers[col_idx]] = val
                else:
                    row_data[headers[col_idx]] = str(val)

        data_count += 1

        # 计算行指纹 + 自然键去重（省去DB查询）
        row_str = json.dumps(row_data, ensure_ascii=False, sort_keys=True)
        row_hash = hashlib.md5(row_str.encode()).hexdigest()

        test_date = ""
        if date_col_idx and date_col_idx - 1 < len(row) and row[date_col_idx - 1] is not None:
            test_date = str(row[date_col_idx - 1])

        natural_key = (sample_name, test_date)
        existing_hash = existing_hashes.get(natural_key) if test_date else None

        sample_id = 0
        if existing_hash is None:
            # 完全新行 → INSERT
            sample_batch.append((
                row_hash, sample_name, test_date, sheet_name, row_str,
                data_count + 1, excel_filename, sheet_name, synced_at
            ))
            existing_hashes[natural_key] = row_hash
            new_rows += 1
        elif existing_hash != row_hash:
            # 行已存在但hash变了（数据被修改）→ UPDATE
            db.conn.execute(
                """UPDATE samples SET row_hash=?, data_json=?, synced_at=?
                   WHERE sample_name=? AND test_date=?""",
                (row_hash, row_str, synced_at, sample_name, test_date)
            )
            db.conn.commit()
            existing_hashes[natural_key] = row_hash
            new_rows += 1  # 计入"有效变化"
            log.info(f"     🔄 更新: {sample_name}({test_date}) hash={row_hash[:8]}")

        # 检查报警
        alarms = check_row_against_specs(
            row_data, sample_name, sheet_specs,
            CONFIG.get("alarm_mode", {})
        )
        for alarm in alarms:
            alarm_batch.append((
                sample_id if sample_id > 0 else 0,
                alarm["sample"], alarm["item"], alarm["value"],
                alarm["spec"], alarm.get("unit", ""), alarm["deviation"],
                alarm["mode"], alarm.get("sheet", sheet_name),
                alarm["mode"], synced_at  # alarm_mode + alarmed_at
            ))

        all_alarms.extend(alarms)

        # 到达批量阈值 → flush
        if len(sample_batch) >= BATCH_SIZE:
            db.insert_samples_batch(sample_batch)
            sample_batch = []
            db.insert_alarms_batch(alarm_batch)
            alarm_batch = []

        if data_count % 10000 == 0:
            log.info(f"     ... 已扫描 {data_count} 行, 报警 {len(all_alarms)} 条, 新增 {new_rows} 行")

    # 刷剩余批次
    if sample_batch:
        db.insert_samples_batch(sample_batch)
    if alarm_batch:
        db.insert_alarms_batch(alarm_batch)

    elapsed = time.time() - t0
    return data_count, all_alarms, new_rows, elapsed


# ─── 飞书通知 ──────────────────────────────────────────
def format_excel_date(val):
    """将Excel序列号日期转为可读格式（46149 → 2026-04-27）"""
    try:
        from datetime import datetime, timedelta
        d = float(val)
        if 40000 < d < 60000:  # Excel日期序列号范围
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=d)).strftime('%m-%d')
    except (ValueError, TypeError, OverflowError):
        pass
    # 如果是 ISO 格式日期，截取年月日
    s = str(val).strip()
    if len(s) >= 10 and s[4] == '-':
        return s[5:10]
    return s[:8]


def send_feishu(alarms, excel_name="", alarm_only=True):
    """发送飞书通知，按样品+检测项聚合展示（只发送 alarm 模式的）"""
    if alarm_only:
        alarms = [a for a in alarms if a.get("mode") == "alarm"]

    if not alarms:
        return

    webhook = CONFIG["feishu"]["webhook_url"]
    if not webhook or webhook.endswith("你的WebHook地址"):
        for a in alarms:
            log.warning(f"  [未发送] {a['sample']} | {a['item']} = {a['value']} (规格:{a['spec']}, 模式:{a['mode']})")
        return

    # 按样品名称 → 检测项 双层聚合
    from collections import defaultdict
    grouped = defaultdict(lambda: defaultdict(list))
    sample_times = {}
    for a in alarms:
        grouped[a["sample"]][a["item"]].append(a)
        # 提取时间
        raw_time = a.get("time", "")
        if raw_time and raw_time not in ("None", ""):
            sample_times[a["sample"]] = format_excel_date(raw_time)

    # 构建聚合展示
    sections = []
    for sample_name in sorted(grouped.keys()):
        items_dict = grouped[sample_name]
        time_str = sample_times.get(sample_name, "")
        time_tag = f"  📅{time_str}" if time_str else ""

        total_items = len(items_dict)
        total_count = sum(len(v) for v in items_dict.values())

        lines = [f"**{sample_name}**{time_tag} — {total_items}项异常"]

        # 按最大偏差排序（从高到低）
        sorted_items = sorted(items_dict.items(),
            key=lambda kv: max(abs(float(a.get("deviation", "0").replace("+","").replace("%","").rstrip("%")) or 0) for a in kv[1]),
            reverse=True
        )

        for item_name, item_alarms in sorted_items:
            count = len(item_alarms)
            # 取偏差最大的那次
            worst = max(item_alarms,
                key=lambda a: abs(float(a.get("deviation", "0").replace("+","").replace("%","").rstrip("%") or 0)))
            val = worst["value"]
            dev = worst["deviation"]
            spec = worst["spec"]
            unit = worst.get("unit", "")

            # 严重度标识
            try:
                dev_pct = abs(float(dev.replace("+","").replace("%","").rstrip("%")))
            except (ValueError, AttributeError):
                dev_pct = 0
            if dev_pct >= 1000:
                tag = "🔴"
            elif dev_pct >= 100:
                tag = "🟠"
            elif dev_pct >= 20:
                tag = "🟡"
            else:
                tag = "⚪"

            repeat = f" 连续{count}次" if count > 1 else ""
            lines.append(f"  {tag} {item_name} = **{val}** {unit} ({spec}) → {dev}{repeat}")

        sections.append("\n".join(lines))

    # 如果内容太长，分段发
    MAX_LENGTH = 1800  # 飞书卡片 markdown 建议长度
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    def send_batch(elements_content, title_suffix=""):
        payload = {
            "msg_type": "interactive",
            "card": {
                "header": {
                    "title": {"tag": "plain_text", "content": f"{CONFIG['feishu']['title']}{title_suffix}"},
                    "template": "red"
                },
                "elements": [
                    {"tag": "markdown", "content": elements_content},
                    {"tag": "note", "elements": [
                        {"tag": "plain_text", "content": f"监控时间: {now_str} | {excel_name}"}
                    ]}
                ]
            }
        }

        try:
            resp = requests.post(webhook, json=payload, timeout=10)
            if resp.status_code == 200:
                log.info(f"✅ 飞书通知已发送 ({len(alarms)} 条)")
            else:
                log.error(f"❌ 发送失败: {resp.status_code}")
        except Exception as e:
            log.error(f"❌ 飞书异常: {e}")

    # 分段发送
    current_batch = []
    current_len = 0
    batch_num = 0
    total_batches = 1

    for section in sections:
        section_len = len(section)
        if current_len + section_len + 2 > MAX_LENGTH and current_batch:
            batch_num += 1
            send_batch("\n\n".join(current_batch), f" ({batch_num}/{total_batches})" if total_batches > 1 else "")
            current_batch = [section]
            current_len = section_len
        else:
            current_batch.append(section)
            current_len += section_len + 2

    if current_batch:
        batch_num += 1
        send_batch("\n\n".join(current_batch), f" ({batch_num}/{total_batches})" if total_batches > 1 else "")

    # 更新总批次标题（如果有多个批次）
    if batch_num > 1:
        total_batches = batch_num


# ─── 写入报警日志（CSV兼容） ──────────────────────────
def write_alarm_log(alarms):
    """将报警记录写入本地CSV文件（向后兼容）"""
    log_path = SCRIPT_DIR / "alarm_history.csv"
    is_new = not log_path.exists()
    with open(log_path, "a", encoding="utf-8") as f:
        if is_new:
            f.write("时间,车间,样品,项目,检测值,规格,偏差,模式\n")
        for a in alarms:
            f.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')},{a.get('sheet','')},{a['sample']},{a['item']},{a['value']},{a['spec']},{a['deviation']},{a['mode']}\n")


# ─── 主扫描（优化版） ──────────────────────────────────────


# ─── 解析Excel文件列表 ────────────────────────────────
def get_excel_files():
    """获取要扫描的Excel文件列表，支持单文件和多文件两种格式
    
    返回: [{"path": str, "password": str|None, "sheets": [], "exclude_sheets": []}]
    """
    excel_cfg = CONFIG["excel"]
    files = excel_cfg.get("files")
    if files and isinstance(files, list) and len(files) > 0:
        return files
    # 向后兼容：单文件格式
    return [{
        "path": excel_cfg.get("path", ""),
        "password": excel_cfg.get("password"),
        "sheets": excel_cfg.get("sheets", []),
        "exclude_sheets": excel_cfg.get("exclude_sheets", [])
    }]


# ─── 启动配置校验 ──────────────────────────────────────
LOCK_FILE = SCRIPT_DIR / ".watchdog.lock"

def check_lock():
    """检查锁文件，防止多实例运行"""
    if LOCK_FILE.exists():
        try:
            pid = int(LOCK_FILE.read_text().strip())
            import errno
            # 跨平台检查进程是否存活
            try:
                os.kill(pid, 0)  # 信号0=仅检查存在性
                log.error(f"❌ 另一个实例已在运行 (PID={pid})，锁文件: {LOCK_FILE}")
                log.error("   如果确定没有其他实例在运行，请删除锁文件后重试")
                sys.exit(1)
            except OSError as e:
                if e.errno == errno.ESRCH:
                    log.warning(f"⚠ 锁文件存在但进程 (PID={pid}) 已不存在，覆盖锁文件")
        except (ValueError, OSError):
            log.warning(f"⚠ 锁文件格式异常，覆盖: {LOCK_FILE}")
    LOCK_FILE.write_text(str(os.getpid()))
    log.info(f"🔒 锁文件已创建 (PID={os.getpid()})")

def release_lock():
    """释放锁文件"""
    if LOCK_FILE.exists():
        try:
            pid = int(LOCK_FILE.read_text().strip())
            if pid == os.getpid():
                LOCK_FILE.unlink()
                log.debug("🔓 锁文件已释放")
        except (ValueError, OSError):
            pass

def validate_config():
    """启动时校验配置完整性"""
    errors = []
    
    # 必须字段
    if "excel" not in CONFIG:
        errors.append("缺少 excel 配置段")
    if "feishu" not in CONFIG:
        errors.append("缺少 feishu 配置段")
    if "database" not in CONFIG:
        errors.append("缺少 database 配置段")
    
    # Excel文件校验
    if not errors:
        excel_files = get_excel_files()
        if not excel_files:
            errors.append("Excel文件列表为空")
        else:
            for i, f in enumerate(excel_files):
                path = f.get("path", "")
                if not path:
                    errors.append(f"files[{i}]: path 为空")
                    continue
                if not os.path.exists(path):
                    errors.append(f"文件不存在: {path}")
    
    # 白名单校验
    whitelist = get_whitelist()
    if whitelist:
        for i, wp in enumerate(whitelist):
            try:
                re.compile(wp)
            except re.error:
                errors.append(f"白名单模式[{i}] '{wp}' 不是有效的正则表达式")
    
    # 飞书Webhook校验
    webhook = CONFIG.get("feishu", {}).get("webhook_url", "")
    if not webhook or webhook.endswith("你的WebHook地址"):
        log.warning("⚠ 飞书Webhook未配置，报警将不会发送到飞书")
    
    if errors:
        log.error("❌ 配置校验失败:")
        for e in errors:
            log.error(f"   - {e}")
        sys.exit(1)
    
    log.info("✅ 配置校验通过")

# ─── 扫描单个Excel文件 ────────────────────────────────
def scan_one_file(file_cfg, existing_hashes, db_specs_dict, db):
    """扫描一个Excel文件中的所有匹配工作表，返回报警列表"""
    path = file_cfg["path"]
    password = file_cfg.get("password")
    target_sheets = file_cfg.get("sheets", [])
    exclude_sheets = file_cfg.get("exclude_sheets", [])
    whitelist_patterns = get_whitelist()
    
    if not os.path.exists(path):
        log.warning("⚠ Excel不存在: " + path)
        return []
    
    # 进度追踪（基于文件路径）
    file_key = path.replace(":", "_").replace("\\", "/").replace("/", "_")
    progress = load_progress()
    mtime = os.path.getmtime(path)
    # 向后兼容：旧格式 {"last_mtime": float} → 新格式 {"last_mtime": {file_key: float}}
    last_mtimes = progress.get("last_mtime", {})
    if isinstance(last_mtimes, (int, float)):
        last_mtimes = {"_old": last_mtimes}
    last_mtime = last_mtimes.get(file_key, 0)
    if mtime <= last_mtime:
        log.info("⏭️  Excel无更新，跳过: " + os.path.basename(path))
        return []
    progress.setdefault("last_mtime", {})
    if isinstance(progress["last_mtime"], (int, float)):
        progress["last_mtime"] = {str(progress["last_mtime"]): progress["last_mtime"]}
    progress["last_mtime"][file_key] = mtime
    save_progress(progress)
    
    wb = open_excel(path, password)
    if wb is None:
        return []
    
    # 过滤工作表
    sheets_to_scan = []
    for sheet_name in wb.sheetnames:
        if target_sheets and sheet_name not in target_sheets:
            continue
        if sheet_name in exclude_sheets:
            continue
        if whitelist_patterns:
            ws = wb[sheet_name]
            if not sheet_has_whitelist_match(ws, sheet_name, whitelist_patterns):
                log.info("⏭️  " + sheet_name + " — 无白名单匹配，跳过")
                continue
        sheets_to_scan.append(sheet_name)
    
    if not sheets_to_scan:
        log.info("⏭️  " + os.path.basename(path) + " — 无工作表需要扫描")
        wb.close()
        return []
    
    log.info("📊 " + os.path.basename(path) + " → 待扫描工作表: " + str(len(sheets_to_scan)) + " 个")
    wb.close()
    
    # 规格缓存
    specs_cache = {}
    if db_specs_dict:
        specs_cache.update(db_specs_dict)
    
    # 并行扫描
    from concurrent.futures import ThreadPoolExecutor, as_completed
    
    all_alarms = []
    total_data = 0
    total_new = 0
    synced_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    def scan_one_sheet(sname):
        try:
            local_db = create_backend(CONFIG.get("database", {"type": "sqlite", "sqlite_path": "./检测数据库.sqlite"}))
            local_db.init_schema()
            wb2 = open_excel(path, password)
            if wb2 is None:
                local_db.close()
                return sname, 0, [], 0, 0
            ws2 = wb2[sname]
            data_count, alarms, new_rows, elapsed = scan_sheet(
                ws2, sname, specs_cache, whitelist_patterns, local_db, existing_hashes
            )
            wb2.close()
            local_db.log_sync(sname, data_count, new_rows, len(alarms), synced_at)
            local_db.close()
            return sname, data_count, alarms, new_rows, elapsed
        except Exception as e:
            log.error("❌ 扫描 " + sname + " 失败: " + str(e))
            return sname, 0, [], 0, 0
    
    MAX_WORKERS = min(4, len(sheets_to_scan))
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(scan_one_sheet, sn): sn for sn in sheets_to_scan}
        for f in as_completed(futures):
            sname, data_count, alarms, new_rows, elapsed = f.result()
            total_data += data_count
            total_new += new_rows
            for a in alarms:
                a["sheet"] = sname
            all_alarms.extend(alarms)
            log.info("  ✅ " + sname + " → " + str(data_count) + "行 | 新增" + str(new_rows) + " | 报警" + str(len(alarms)) + " | " + f"{elapsed:.2f}s")
            if alarms:
                for msg in ["     " + a["sample"] + " " + a["item"] + " = " + str(a["value"]) + " " + a.get("unit","") + " (" + a["spec"] + ") [" + a["mode"] + "]" for a in alarms[:5]]:
                    log.info(msg)
                if len(alarms) > 5:
                    log.info("     ... 共 " + str(len(alarms)) + " 条")
    
    if total_new > 0:
        log.info("  💾 " + os.path.basename(path) + ": 新增" + str(total_new) + "行, 报警" + str(len(all_alarms)) + "条")
    return all_alarms


# ─── 主扫描（支持多文件） ─────────────────────────────
def scan_all(db):
    """扫描所有Excel文件的所有工作表，返回聚合报警列表"""
    
    excel_files = get_excel_files()
    whitelist_patterns = get_whitelist()
    
    if not excel_files:
        log.warning("⚠ 未配置任何Excel文件")
        return []
    
    # 预加载：内存hash + DB规格缓存
    existing_hashes = db.get_all_existing_hashes()
    db_specs_dict = db.get_specs_dict()
    log.info("📦 已缓存 " + str(len(existing_hashes)) + " 个行指纹, 规格缓存就绪")
    
    all_alarms = []
    for file_cfg in excel_files:
        if STOP_FLAG:
            break
        alarms = scan_one_file(file_cfg, existing_hashes, db_specs_dict, db)
        all_alarms.extend(alarms)
    
    log.info("📊 总计: " + str(len(all_alarms)) + " 条报警 (来自 " + str(len(excel_files)) + " 个Excel文件)")
    if all_alarms:
        log.info("💾 数据库: samples=" + str(db.get_sample_count()) + ", alarm_log=" + str(db.get_alarm_count()))
    return all_alarms

# ─── 报警去重 ──────────────────────────────────────────
LAST_ALARM_VALUES = {}  # {(sample, item): (value, time)} 用于去重

def dedup_alarms(alarms):
    """过滤掉同一样品同一项目与上次扫描值不变的报警"""
    if not alarms:
        return []
    deduped = []
    for a in alarms:
        key = (a.get("sample", ""), a.get("item", ""))
        new_val = a.get("value")
        old_val = LAST_ALARM_VALUES.get(key)
        if old_val is None or abs(float(new_val) - float(old_val[0])) > 0.0001:
            deduped.append(a)
            LAST_ALARM_VALUES[key] = (new_val, time.time())
    return deduped

def main():
    once = "--once" in sys.argv
    daemon_child = "--daemon-child" in sys.argv
    daemon = "--daemon" in sys.argv
    init_db_only = "--init-db" in sys.argv
    skip_validation = "--skip-validation" in sys.argv

    # 后台模式：先 fork 子进程，父进程立即退出
    if daemon:
        import subprocess
        script = Path(__file__).resolve()
        cmd = [sys.executable, str(script), "--daemon-child"]
        log_path = SCRIPT_DIR / "watchdog_daemon.log"
        flags = subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0
        with open(log_path, "w") as lf:
            proc = subprocess.Popen(cmd, stdout=lf, stderr=lf, creationflags=flags)
        print(f"🔄 已后台启动 (PID: {proc.pid})")
        print(f"📋 日志: {log_path}")
        return

    # 校验配置 + 锁文件
    if not skip_validation:
        validate_config()
    compile_whitelist()
    check_lock()
    import atexit
    atexit.register(release_lock)

    # 创建数据库连接
    db_config = CONFIG.get("database", {"type": "sqlite", "sqlite_path": "./检测数据库.sqlite"})
    db = create_backend(db_config)

    if not db.is_connected():
        log.error("❌ 无法连接数据库，请检查 config.json 中的 database 配置")
        sys.exit(1)

    # 初始化表结构
    db.init_schema()
    log.info("✅ 数据库表结构已就绪")
    log.info(f"📊 当前: samples={db.get_sample_count()} | alarm_log={db.get_alarm_count()} | specs={len(db.get_specs())}")

    # 仅初始化数据库
    if init_db_only:
        log.info("✅ --init-db 完成")
        db.close()
        return

    # 正常模式（前台的持续监控或 --once）
    log.info("=" * 50)
    log.info("🔍 中控检测数据监控报警器 V3.1")
    excel_count = len(get_excel_files())
    log.info(f"   Excel文件数: {excel_count}")
    for ef in get_excel_files()[:3]:
        log.info(f"     · {ef.get('path', '?')}")
    if excel_count > 3:
        log.info(f"     ... 共 {excel_count} 个文件")
    log.info(f"   数据库: {db_config.get('type', 'sqlite')}")
    if get_whitelist():
        log.info(f"   白名单: {len(get_whitelist())} 条模式 (已启用)")
    log.info("=" * 50)

    # 一次模式
    if once:
        alarms = scan_all(db)
        alarms = dedup_alarms(alarms)
        if alarms:
            send_feishu(alarms, "单次扫描")
            write_alarm_log(alarms)
            monitor = [a for a in alarms if a.get("mode") == "monitor"]
            if monitor:
                log.info(f"📝 Monitor记录: {len(monitor)} 条（仅记录，未通知）")
        else:
            log.info("✅ 无报警")
        db.close()
        return

    # 持续监控模式
    excel_files = get_excel_files()
    log.info(f"📂 监控 {len(excel_files)} 个Excel文件:")
    for f in excel_files:
        log.info(f"   · {f.get('path', '?')}")
    
    health_interval = CONFIG.get("watch", {}).get("health_check_interval", 10)
    health_counter = 0
    while not STOP_FLAG:
        try:
            alarms = scan_all(db)
            alarms = dedup_alarms(alarms)
            if alarms:
                send_feishu(alarms, "多文件监控")
                write_alarm_log(alarms)
                monitor = [a for a in alarms if a.get("mode") == "monitor"]
                if monitor:
                    log.info(f"📝 Monitor记录: {len(monitor)} 条（仅记录，未通知）")
            else:
                log.debug("✅ 无报警")

            # 定期DB健康检查
            health_counter += 1
            if health_counter >= health_interval:
                try:
                    db.get_sample_count()
                except Exception as e:
                    log.error(f"❌ DB连接异常，尝试重连: {e}")
                    db.close()
                    db_config = CONFIG.get("database", {"type": "sqlite", "sqlite_path": "./检测数据库.sqlite"})
                    db = create_backend(db_config)
                    db.init_schema()
                    log.info("✅ DB重连成功")
                health_counter = 0

        except Exception as e:
            log.error(f"❌ 执行异常: {e}", exc_info=True)

        # 分段睡眠（支持快速响应STOP信号）
        sleep_interval = CONFIG["watch"]["interval_seconds"]
        for _ in range(min(sleep_interval, 10)):
            if STOP_FLAG:
                break
            time.sleep(1)

    log.info("🛑 监控已停止")
    db.close()

if __name__ == "__main__":
    main()
